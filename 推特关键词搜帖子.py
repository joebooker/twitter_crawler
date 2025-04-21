#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :csxq_keyword_search1.2_modified.py

import os
import re
import json
import time
import requests
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers


class CsxqTwitterKeywordSearch:
    def __init__(self, cookies):
        self.headers = {
            "authority": "twitter.com",
            "accept": "*/*",
            "accept-language": "zh-CN,zh;q=0.9",
            "authorization": "Bearer AAAAAAAAAAAAAAAAAAAAANRILgAAAAAAnNwIzUejRCOuH5E6I8xnZz4puTs%3D1Zv7ttfk8LF81IUq16cHjhLTvJu4FA33AGWWjCpTnA",
            "content-type": "application/json",
            "referer": "https://twitter.com/search?f=top&q=(from%3Anba)&src=typed_query",
            "sec-ch-ua": "\"Not/A)Brand\";v=\"99\", \"Google Chrome\";v=\"115\", \"Chromium\";v=\"115\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"macOS\"",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            "x-client-transaction-id": "qksff6z6OrIU9haO5hguxhlTVWPrEXzbvZ4KfRXEeaBw7HGViMK2vIQP8adWBxC7mKK2N6pJ0aOWxlRZwwpluYThescmqw",
            "x-client-uuid": "436600b9-cde7-4dac-8ef8-cf411d6ed659",
            "x-csrf-token": "",
            "x-twitter-active-user": "yes",
            "x-twitter-auth-type": "OAuth2Session",
            "x-twitter-client-language": "en"
        }
        self.cookies_list = [self.cookie_str_to_dict(cookie) for cookie in cookies]
        self.current_cookie_index = 0
        self.csv_file_paths = {}  # 用于存储每个关键词对应的CSV文件路径
        self.xlsx_file_paths = {}  # 用于存储每个关键词对应的XLSX文件路径

    def cookie_str_to_dict(self, cookie_str: str) -> dict:
        cookie_dict = {}
        cookies = cookie_str.split('; ')
        for cookie in cookies:
            if '=' in cookie:
                key, value = cookie.split('=', 1)
                cookie_dict[key] = value
        return cookie_dict

    def get_params(self, cursor, searchCondition,search_type):
        variables = {
            "rawQuery": searchCondition,
            "count": 20,
            "querySource": "typed_query",
            "product": search_type
        }
        if cursor:
            variables["cursor"] = cursor
        params = {
            "variables": json.dumps(variables, separators=(",", ":")),
            "features": "{\"rweb_tipjar_consumption_enabled\":true,\"responsive_web_graphql_exclude_directive_enabled\":true,"
                        "\"verified_phone_label_enabled\":false,\"creator_subscriptions_tweet_preview_api_enabled\":true,"
                        "\"responsive_web_graphql_timeline_navigation_enabled\":true,\"responsive_web_graphql_skip_user_profile_image_extensions_enabled\":false,"
                        "\"communities_web_enable_tweet_community_results_fetch\":true,\"c9s_tweet_anatomy_moderator_badge_enabled\":true,"
                        "\"articles_preview_enabled\":true,\"tweetypie_unmention_optimization_enabled\":true,\"responsive_web_edit_tweet_api_enabled\":true,"
                        "\"graphql_is_translatable_rweb_tweet_is_translatable_enabled\":true,\"view_counts_everywhere_api_enabled\":true,"
                        "\"longform_notetweets_consumption_enabled\":true,\"responsive_web_twitter_article_tweet_consumption_enabled\":true,"
                        "\"tweet_awards_web_tipping_enabled\":false,\"creator_subscriptions_quote_tweet_preview_enabled\":false,"
                        "\"freedom_of_speech_not_reach_fetch_enabled\":true,\"standardized_nudges_misinfo\":true,"
                        "\"tweet_with_visibility_results_prefer_gql_limited_actions_policy_enabled\":true,"
                        "\"rweb_video_timestamps_enabled\":true,\"longform_notetweets_rich_text_read_enabled\":true,"
                        "\"longform_notetweets_inline_media_enabled\":true,\"responsive_web_enhance_cards_enabled\":false}"
        }
        return params

    def get(self, cursor, searchCondition,search_type):
        while True:
            self.headers["x-csrf-token"] = self.cookies_list[self.current_cookie_index].get('ct0', '')
            cookies = self.cookies_list[self.current_cookie_index]
            print(f"使用的cookie索引: {self.current_cookie_index}, 类型: {type(cookies)}")
            url = "https://x.com/i/api/graphql/6uoFezW1o4e-n-VI5vfksA/SearchTimeline"
            params = self.get_params(cursor, searchCondition,search_type)
            try:
                response = requests.get(url, headers=self.headers, cookies=cookies, params=params, timeout=(3, 10))
                print(f"响应状态码：{response.status_code}")
                # print(response.headers)  # 根据需要开启

                if response.status_code == 429:  # 超过速率限制
                    print("超过速率限制。切换到下一个cookie。")
                    self.current_cookie_index = (self.current_cookie_index + 1) % len(self.cookies_list)
                    print(f"新的cookie索引: {self.current_cookie_index}")
                    print(self.cookies_list[self.current_cookie_index])
                    time.sleep(20)  # 等待一段时间后重试
                    continue
                if response.status_code in [401, 403]:
                    print("访问被拒绝。切换到下一个cookie。")
                    self.current_cookie_index = (self.current_cookie_index + 1) % len(self.cookies_list)
                    print(f"新的cookie索引: {self.current_cookie_index}")
                    time.sleep(10)
                    continue

                if response.status_code == 200:
                    data = response.json()
                    return data
                else:
                    print(f"请求失败，状态码：{response.status_code}")
                    return None
            except Exception as e:
                print(f"查询搜索API时出错: {e}")
                time.sleep(10)  # 等待一段时间后重试

    def contains_non_english_characters(self, text):
        # 正则表达式：匹配任何非英文字符（包括中文、特殊符号等）
        # 允许字母、数字、空格、常见标点符号和表情符号（emoji）
        ssss = r'[^\x00-\x7F\s.,!?\'"’@-]'
        try:
            is_text_emojis = re.search(r'[\U0001F600-\U0001F64F\U0001F300-\U0001F9FF]', text)
        except re.error:
            # 如果text中包含未定义的emoji，忽略
            is_text_emojis = None
        if is_text_emojis:
            # 去掉原文本中的表情符号
            text_without_emojis = re.sub(r'[\U0001F600-\U0001F64F\U0001F300-\U0001F9FF]', '', text)
            # 检查是否包含非英文字符（中文、日文等），如果存在，返回True
            return bool(re.search(ssss, text_without_emojis))
        else:
            return bool(re.search(ssss, text))

    def parse_data(self, entries, keyword):
        resultList = []
        earliest_date = None  # 用于记录最早日期

        def transTime(dd):
            GMT_FORMAT = '%a %b %d %H:%M:%S +0000 %Y'
            timeArray = datetime.datetime.strptime(dd, GMT_FORMAT)
            return timeArray.strftime("%Y-%m-%d %H:%M:%S")

        contentList = []
        for ent in entries:
            try:
                entryId = ent.get('entryId', "")
                if 'tweet' in entryId:
                    l_result = ent['content']['itemContent']['tweet_results']['result'] if ent['content'].get(
                        'itemContent') else None
                    if l_result:
                        contentList.append(l_result)
                elif "profile-conversation" in entryId:
                    items = ent['content']['items']
                    for i in items:
                        l_result = i['item']['itemContent']['tweet_results']['result'] if i['item'].get(
                            'itemContent') else None
                        if l_result:
                            contentList.append(l_result)
            except Exception as e:
                print(f"处理条目时出错: {e}")
                continue

        for l in contentList:
            try:
                result = l.get('tweet') if l.get('tweet') else l
                legacy = result['legacy']
                id_str = legacy.get('id_str')
                core = result['core']
                text_created_at = transTime(legacy.get('created_at'))
                tweet_datetime = datetime.datetime.strptime(text_created_at, "%Y-%m-%d %H:%M:%S")

                # 更新最早日期
                if earliest_date is None or tweet_datetime < earliest_date:
                    earliest_date = tweet_datetime

                user_legacy_data = core.get('user_results', {}).get('result', {}).get('legacy', {})
                count_created_at = transTime(user_legacy_data.get('created_at')) if user_legacy_data.get(
                    'created_at') else None
                is_verified = core['user_results']['result'].get('legacy', {}).get('verified')
                blue_verified = core['user_results']['result'].get('is_blue_verified')
                has_graduated_access = core['user_results']['result'].get('has_graduated_access')
                is_retweet = legacy.get('retweeted')
                professional_type = core['user_results']['result'].get('professional', {}).get('professional_type')
                full_text = legacy.get('full_text')
                # 捕获额外的指标
                note_tweet = result.get('note_tweet')
                favorite_count = legacy.get('favorite_count')
                reply_count = legacy.get('reply_count')
                retweet_count = legacy.get('retweet_count', 0)
                quote_count = legacy.get('quote_count', 0)
                retweet_count += quote_count
                if note_tweet:
                    try:
                        full_text = note_tweet['note_tweet_results']['result']['text']
                    except:
                        pass
                # 移除英文过滤
                # if self.contains_non_english_characters(full_text):
                #     continue
                reply_id = legacy.get('in_reply_to_status_id_str')
                is_reply = "回复" if reply_id else "原始推文"
                professional = core['user_results']['result'].get("professional")
                is_company = "是" if professional else "否"
                u_legacy = core['user_results']['result']['legacy']
                hash_uname = u_legacy.get('screen_name')
                description = u_legacy['description']
                friends_count = u_legacy['friends_count']
                followers_count = u_legacy.get('followers_count')
                statuses_count = u_legacy.get('statuses_count')
                location = u_legacy.get('location')
                url = f'https://x.com/{hash_uname}/status/{id_str}'
                item = {
                    "帖子id": id_str,
                    '网址': url,
                    "内容": full_text,
                    "是否回复": is_reply,
                    "发布时间": text_created_at,
                    "账号创建时间": count_created_at,
                    "点赞": favorite_count,
                    "评论": reply_count,
                    "转发": retweet_count,
                    "用户名": hash_uname,
                    "简介": description,
                    "ip": location,
                    "is_professional": is_company,
                    "是否转帖": is_retweet,
                    "帖子数量": statuses_count,
                    "粉丝量": followers_count,
                    "关注量": friends_count,
                    "关键词": keyword,
                    "is_verified": is_verified,
                    "is_blue_verified": blue_verified,
                    "has_graduated_access": has_graduated_access,
                    "professional_type": professional_type
                }
                print(item)
                resultList.append(item)
            except Exception as e:
                print(f"解析推文数据时出错: {e}")

        return resultList, earliest_date  # 返回最早日期

    def get_cursor(self, dataJson):
        instructions = dataJson.get('data', {}).get('search_by_raw_query', {}).get('search_timeline', {}).get(
            'timeline', {}).get('instructions', [])
        entries = []
        cursor = None
        for ins in instructions:
            if ins.get('type') == "TimelineAddEntries":
                entries = ins.get('entries', [])
                for ent in entries:
                    content = ent.get('content', {})
                    cursorType = content.get('cursorType')
                    if cursorType == 'Bottom':
                        cursor = content.get('value')
                        break
        if not cursor:
            for ins in instructions:
                entry = ins.get('entry', {})
                content = entry.get('content', {})
                cursorType = content.get('cursorType')
                if cursorType == 'Bottom':
                    cursor = content.get('value')
                    break
        return cursor, entries

    def run(self, word, start_date, end_date,language,search_type):
        all_scraped = False  # 标记是否已全部爬取
        current_end_date = end_date

        while not all_scraped:

            search_condition = f"{word} lang:{language} until:{current_end_date} since:{start_date}"
            cursor = ""
            print(f"开始爬取关键词: {word}，时间区间: {start_date} 到 {current_end_date}")

            try:
                batch_earliest_date = None  # 记录本次外层循环中的最早日期
                while True:
                    print(f"正在爬取关键词: {word}, 时间区间: {start_date} 到 {current_end_date}, cursor: {cursor}")
                    resqJson = self.get(cursor, search_condition,search_type)
                    if not resqJson:
                        print("未获取到数据，可能是API变化或其他问题。")
                        break

                    cursor, entries = self.get_cursor(resqJson)
                    if entries:
                        resultList, batch_date = self.parse_data(entries, search_condition)
                        self.save_data_csv(resultList, word)

                        # 更新本次外层循环的最早日期
                        if batch_date:
                            if batch_earliest_date is None or batch_date < batch_earliest_date:
                                batch_earliest_date = batch_date
                    else:
                        print("没有更多的条目，停止爬取。")
                        break

                if batch_earliest_date:
                    print(f"本次批次最早日期: {batch_earliest_date.strftime('%Y-%m-%d %H:%M:%S')}")
                    # 检查是否已经覆盖到起始日期
                    if batch_earliest_date <= datetime.datetime.strptime(start_date, "%Y-%m-%d"):
                        all_scraped = True
                        print(f"已覆盖所有指定时间段的数据。")
                        break

                    # 更新时间区间以继续爬取
                    new_end_date = (batch_earliest_date - datetime.timedelta(seconds=1)).strftime("%Y-%m-%d")
                    print(f"调整时间区间: {start_date} 到 {new_end_date}，继续爬取")
                    current_end_date = new_end_date
                else:
                    print("未记录到最早日期，可能数据有误，停止爬取。")
                    break

            except Exception as e:
                print(f"发生错误：{e}")
                # 在发生异常时，尝试继续下一时间段
                if 'batch_earliest_date' in locals() and batch_earliest_date:
                    new_end_date = (batch_earliest_date - datetime.timedelta(seconds=1)).strftime("%Y-%m-%d")
                    print(f"由于错误，调整时间区间: {start_date} 到 {new_end_date}，继续爬取")
                    current_end_date = new_end_date
                else:
                    print("无法确定新的结束日期，退出爬取。")
                    break

            # 延迟以避免触发速率限制
            time.sleep(0.5)

    def save_data_csv(self, resultList, keyword):
        """
        将数据追加保存到 CSV 文件中。
        """
        if not resultList:
            return

        df = pd.DataFrame(resultList)
        # 初始化文件路径
        if keyword not in self.csv_file_paths:
            self.csv_file_paths[keyword] = f'./{keyword}.csv'
            self.xlsx_file_paths[keyword] = f'./{keyword}.xlsx'

        file_path = self.csv_file_paths[keyword]

        if not os.path.exists(file_path):
            # 如果文件不存在，写入标题
            df.to_csv(file_path, index=False, mode='w', encoding='utf-8-sig')
        else:
            # 如果文件已存在，追加写入，不写入标题
            df.to_csv(file_path, index=False, mode='a', header=False, encoding='utf-8-sig')

        print(f"已保存 {len(resultList)} 条数据到 CSV 文件: {file_path}")

    def convert_csv_to_xlsx(self, keyword):
        """
        将指定关键词的 CSV 文件转换为 XLSX 文件，并格式化 ID 列为文本格式。
        """
        csv_path = self.csv_file_paths.get(keyword)
        xlsx_path = self.xlsx_file_paths.get(keyword)

        if not csv_path or not os.path.exists(csv_path):
            print(f"CSV 文件不存在，无法转换: {csv_path}")
            return

        # 读取 CSV 文件时，将 '帖子id' 列作为字符串读取以保留格式
        try:
            df = pd.read_csv(csv_path, dtype={'帖子id': str}, encoding='utf-8-sig')
        except Exception as e:
            print(f"读取 CSV 文件失败: {e}")
            return

        # 使用 Pandas 将 DataFrame 写入 Excel 文件
        try:
            df.to_excel(xlsx_path, index=False, sheet_name='Sheet1')
        except Exception as e:
            print(f"写入 Excel 文件失败: {e}")
            return

        # 使用 openpyxl 打开 Excel 文件并格式化 ID 列
        try:
            wb = load_workbook(xlsx_path)
            ws = wb['Sheet1']

            # 获取 '帖子id' 列的索引
            id_column = '帖子id'
            if id_column not in df.columns:
                print(f"列 '{id_column}' 不存在于 CSV 文件中。")
                return

            id_col_idx = df.columns.get_loc(id_column) + 1  # openpyxl 是1-based

            # 设置整个 '帖子id' 列的数字格式为文本
            for cell in ws.iter_cols(min_col=id_col_idx, max_col=id_col_idx):
                for c in cell:
                    c.number_format = numbers.FORMAT_TEXT

            wb.save(xlsx_path)
            print(f"成功将 CSV 转换为 XLSX 并格式化 '{id_column}' 列: {xlsx_path}")
        except Exception as e:
            print(f"格式化 Excel 文件失败: {e}")

    def convert_all_csv_to_xlsx(self):
        """
        将所有关键词的 CSV 文件转换为 XLSX 文件。
        """
        for keyword in self.csv_file_paths:
            print(f"转换关键词 '{keyword}' 的 CSV 文件为 XLSX 文件。")
            self.convert_csv_to_xlsx(keyword)

    def main(self, start_date, end_date,wordList,language,search_type):
        for word in wordList:
            print(f"开始处理关键词: {word}")
            self.run(word, start_date, end_date,language,search_type)
            print("等待15秒...")
            time.sleep(15)

        print("所有爬取任务完成，开始转换 CSV 到 XLSX。")
        self.convert_all_csv_to_xlsx()
        print("转换完成。")


if __name__ == '__main__':
    cookie_strings =[
        "night_mode=2; kdt=ZF58zUioqAhS8neVIUroJfgpvRNgRTlV3PaCHNU4; dnt=1; guest_id=v1%3A173150011582581354; guest_id_marketing=v1%3A173150011582581354; guest_id_ads=v1%3A173150011582581354; auth_token=a61110c56ce1fc9c47c2ab3b4d59ff7bb01134f2; ct0=8e22435c81030160ef1de3cdff3309608524a14945df0d6387cd20986b6985d17ecf3adf29cc09ad1827fc7d74199a2900a70a929c26597195300da1dc24b2bb11e7f2264110aa1a6fd5c4ebea6a5741; twid=u%3D1787472284401487872;lang=en",
        'guest_id=173760735665435804; night_mode=2; guest_id_marketing=v1%3A173760735665435804; guest_id_ads=v1%3A173760735665435804; gt=1882287801905062199; kdt=QOpuADsjI7wXWtywuJcuH6m1nD5sVJg9opK9VPaR; auth_token=02be8f92a91e883ee6dc7fb3ebf01f29d2d162f2; ct0=32219b2481da90976bbd9088dc6f07d7f326684ef45abb56e718a815a162f64e98049f367e201204af8251ba310dea1d12419c3f883d91f9d90aa6d52ae4ce10fd3dc41ba4c26827cf316660ef0eba65; att=1-NGQAmQbZKDYG1uckT3SgBFJ76ld7LJK4VJ4WMapS; lang=en; twid=u%3D1862824059429679104;',
        'guest_id=173760741898563219; night_mode=2; guest_id_marketing=v1%3A173760741898563219; guest_id_ads=v1%3A173760741898563219; gt=1882288063700959669; kdt=iAaStGVwarIW0koEmAs7ROnqWCwJcLNoS4fwqtcw; auth_token=213d7eb84c466c77a402e0a18da74718e279955a; ct0=8263d1b9aa51dc776b65e3bf0aedb02aa29760c9b8d710da3d8bd58007cc87e23b7dd31ec569c4209854bfaca984d49740278553576f531964da8d56af2baa18661bd185e1fd392b1fb684d4f103bfad; lang=en; twid=u%3D1863294481212260352;',
    ]
    ctks = CsxqTwitterKeywordSearch(cookie_strings)
    start_date = '2024-07-06'
    end_date = '2024-12-27'
    wordList = ["climate change"]#多个关键词以逗号分开
    language='en'#语言，英文：en 中文：zh
    search_type = 'Latest'#Latest 和 Top
    ctks.main(start_date, end_date,wordList,language,search_type)
